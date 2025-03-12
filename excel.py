from os import getcwd
from tkinter import Tk, filedialog
from chase import chase
from openpyxl import Workbook, load_workbook
from citi import citi
from american import american_express
from aspire import aspire
import us

from salliemae import salliemae


path=getcwd()

def tkinter_window():
	root=Tk()
	root.title("Choose Your Excel File")
	root.filename=filedialog.askopenfilename(initialdir=path, title="Select A File")
	root.destroy()
	return root.filename
# Create a new workbook
# wb = Workbook()
# ws = wb.active

# # Write data to cells
# ws['A1'] = 'Name'
# ws['B1'] = 'Age'
# ws['A2'] = 'John'
# ws['B2'] = 30

# # Save the workbook
# wb.save('example.xlsx')

# Load an existing workbook
path = tkinter_window()
wb = load_workbook(path)
ws = wb.active

# Read data from cells
# name = ws['A2'].value
# age = ws['B2'].value
# print(f"Name: {name}, Age: {age}")

# Iterate through rows
# ws['G2'] = '512334241'
# ws['G3'] = '619405542'
# ws['D2'] = 'Trythisout4200@outlook.com'
# ws['D3'] = 'pytcli21@outlook.com'
# ws['I2'] = 'Lawrence Township'
# ws['K2'] = '08648'
# ws['J2'] = 'New Jersey'
# wb.save('dummy_data.xlsx')
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    lastname = row[1]
    address = row[2]
    email = row[3]
    phone = row[4]
    dob = row[5]
    ssn = f'{row[6]}'
    annual_income = row[7]
    city = row[8]
    state = row[9]
    zipcode = row[10]
    coverage_addendum = bool(row[11])
    credit_protection = bool(row[12])
    consent = bool(row[13])
    mothers_name = row[14]
    residence_own = row[15]
    primary_income = row[16]
    degree_type = row[17]
    uni_state=row[18]
    uni_name=row[19].upper()
    uni_zip=row[20]
    stream=row[21]
    year_of_degree=row[22]
    time_attending=row[23]
    school_time_period=row[24]
    loan_start_time=row[25].split('/')
    loan_end_time=row[26].split('/')
    graduation_time=row[27].split('/')
    financial_aid=row[28]
    loan_amount=row[29]
    office_phone=row[30]
    username=row[31]
    password=row[32]
    sdob=dob.split('/')
    # print(primary_income)
    # delta(
    #     first_name=name,
    #     last_name=lastname,
    #     email=email,
    #     username=username,
    #     password="Pikachu123",
    #     confirm_password="Pikachu123",
    #     dob=dob,
    #     address=address,
    #     city=city,
    #     state=state,
    #     zip_code=zipcode,
    #     phone_number=phone
    # )
    citi(
        firstname=name,
        lastname=lastname,
        ssn=ssn,
        dob=dob,
        address=address,
        zip=zipcode,
        phone=phone,
        email=email,
        income=annual_income
    )
    american_express(
        firstname=name,
        lastname=lastname,
        ssn=ssn,
        dob=dob,
        address=address,
        zip=zipcode,
        phone=phone,
        email=email,
        income=annual_income,
        source=primary_income,
        # state=state,
        # city=city
    )
    chase(
        firstname=name,
        lastname=lastname,
        mother_maiden_name=mothers_name,
        address=address,
        city=city,
        state=f'{us.states.lookup(state).abbr}',
        zip=zipcode,
        email=email,
        phone=phone,
        dob=dob,
        ssn=ssn,
        residence=residence_own,
        source=primary_income,
        income=annual_income
    )
    aspire(
        firstname=name,
        lastname=lastname,
        address=address,
        city=city,
        state= f'{us.states.lookup(state).abbr} - {state}',
        zip=zipcode,
        email=email,
        phone=phone,
        dob=dob,
        ssn=ssn,
        annual_income=annual_income,
        coverage_addendum=False,
        credit_protection=False,
        consent=consent
    )
    salliemae(
        first_name=name,
        last_name=lastname,
        email_address=email,
        phone_number=phone,
        birth_month=sdob[0],
        birth_day=sdob[1],
        birth_year=sdob[2],
        ssn_number=ssn,
        street_address=address,
        city_name=city,
        degree_type=degree_type,
        state_code_=f'{us.states.lookup(state).abbr}',
        zip_code_=zipcode,
        school_state_=uni_state,
        school_name_=uni_name,
        school_zip_code=f"{uni_zip}",
        target_stream=stream,
        year_of_degree=year_of_degree,
        time_attending=time_attending,
        school_time_period=school_time_period,
        graduation_month=graduation_time[0],
        graduation_year=graduation_time[1],
        financial_aid_amount=financial_aid,
        loan_asking_amount=loan_amount,
        employement_type=primary_income.capitalize(),
        annual_income=annual_income,
        office_phone=office_phone,
        username_=username,
        password_=password
    )


# Add a new sheet
# new_sheet = wb.create_sheet("Sheet2")

# # Save changes
# wb.save('example.xlsx')